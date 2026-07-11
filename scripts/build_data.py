"""
Extrae y normaliza el Excel fuente (4 pestañas) en un único JSON de datos
para la plataforma. Ejecutar: python3 scripts/build_data.py
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "source" / "agenda-corte.xlsx"
OUT = ROOT / "src" / "data" / "directorio.json"

HONORIFICOS = re.compile(r"^(SRTA|SRA|SR|OF|DON|DO[ÑN]A)\.?\s*", re.IGNORECASE)
ACRONIMOS = {"Capj": "CAPJ", "Ica": "ICA", "Cinj": "CINJ", "Csmp": "CSMP", "Ule": "ULE", "Coccmp": "COCCMP"}


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_key(s: str) -> str:
    if not s:
        return ""
    s = strip_accents(str(s)).upper()
    s = re.sub(r"[°º.]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def title_case(name: str) -> str:
    if not name:
        return name

    def cap_word(w: str) -> str:
        return "-".join(s[:1].upper() + s[1:].lower() if s else s for s in w.split("-"))

    words = re.sub(r"\s+", " ", name.strip()).split(" ")
    out = [cap_word(w) for w in words]
    out = [ACRONIMOS.get(w, w) for w in out]
    return " ".join(out)


def sort_emails(emails: list[str]) -> list[str]:
    return sorted(emails, key=lambda e: 0 if e.endswith("@pjud.cl") else 1)


COMUNAS = [
    "La Serena",
    "Coquimbo",
    "Ovalle",
    "Illapel",
    "Vicuña",
    "Combarbalá",
    "Los Vilos",
    "Andacollo",
]


def extraer_comuna(nombre: str) -> str | None:
    k = norm_key(nombre)
    for comuna in COMUNAS:
        if norm_key(comuna) in k:
            return comuna
    return None


def clean_name(raw: str) -> tuple[str, bool]:
    """Devuelve (nombre_legible, es_generico)."""
    if raw is None:
        return "", False
    raw = str(raw).strip()
    if not raw:
        return "", False
    if norm_key(raw) in {"CORREO GENERAL", "CORREO GENERAL UNIDAD"}:
        return "Casilla general", True
    if norm_key(raw) == "VACANTE":
        return "", False
    cleaned = HONORIFICOS.sub("", raw).strip()
    cleaned = re.sub(r"\s*\(.*?\)\s*", " ", cleaned).strip()
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return title_case(cleaned), False


def split_emails(raw) -> list[str]:
    if raw is None:
        return []
    raw = str(raw)
    parts = re.split(r"[;/,]| y ", raw)
    out = []
    for p in parts:
        p = p.strip()
        m = re.search(r"[\w.\-+]+@[\w.\-]+\.[a-zA-Z]{2,}", p)
        if m:
            out.append(m.group(0).lower())
    seen = set()
    uniq = []
    for e in out:
        if e not in seen:
            seen.add(e)
            uniq.append(e)
    return uniq


def slugify(s: str) -> str:
    s = strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------------------------------------------------------------------
# 1) Pestaña "Anexos": directorio interno de la Corte con anexo telefónico
# ---------------------------------------------------------------------------
ws = wb["Anexos"]
SECCION_ANEXOS = {
    "INSOLVENCIA": "insolvencia",
    "CINJ": "cinj",
    "COCCMP": "csmp",
}
UNIDAD_DISPLAY_ANEXOS = {
    "CORTE DE APELACIONES LA SERENA": "Contacto General",
    "ATENCION DE PUBLICO": "Atención de Público",
    "PRESIDENCIA": "Presidencia",
    "MINISTROS": "Ministros",
    "SECRETARIA": "Secretaría",
    "ADMINISTRADORA": "Administración",
    "1 FISCALIA": "1ª Fiscalía Judicial",
    "2 FISCALIA": "2ª Fiscalía Judicial",
    "SALAS": "Salas",
    "SALA DE REUNIONES": "Sala de Reuniones",
    "RELATORES": "Relatoría",
    "UNIDAD DE CAUSAS Y SALAS": "Unidad de Causas y Salas",
    "OFICIAL DE SALA": "Oficial de Sala",
    "ASISTENTE PDTE": "Asistencia de Presidencia",
    "UNIDAD DE SERVICIOS": "Unidad de Servicios",
    "UNIDAD DE CONTROL DE GESTION Y COORD CON TRINUNALES DE 1A INSTANCIA": "Unidad de Control de Gestión y Coordinación con Tribunales",
    "BIBLIOTECA": "Biblioteca",
    "UNIDAD DE PLENO": "Unidad de Pleno",
    "OFICINA DERECHOS HUMANOS": "Oficina de Derechos Humanos",
    "PERIODISTAS": "Comunicaciones y Prensa",
    "GENDARMERIA": "Gendarmería (enlace)",
    "PORTERIA": "Portería y Seguridad",
}


def display_unidad_anexos(raw: str) -> str:
    return UNIDAD_DISPLAY_ANEXOS.get(norm_key(raw), title_case(raw))
anexos_por_correo: dict[str, dict] = {}
anexos_rows = []
for r in range(3, 88):
    unidad = ws.cell(row=r, column=2).value
    nombre_raw = ws.cell(row=r, column=3).value
    correo_raw = ws.cell(row=r, column=4).value
    anexo = ws.cell(row=r, column=5).value
    cumple = ws.cell(row=r, column=6).value
    if not unidad and not nombre_raw:
        continue
    unidad = str(unidad).strip() if unidad else ""
    unidad_display = display_unidad_anexos(unidad)
    es_vacante = nombre_raw is not None and norm_key(str(nombre_raw)) == "VACANTE"
    nombre, es_generico = clean_name(nombre_raw)
    if nombre.startswith("Contacto "):
        es_generico = True
    correos = split_emails(correo_raw)
    seccion = SECCION_ANEXOS.get(norm_key(unidad), "corte")
    row = {
        "unidad": unidad_display,
        "nombre": nombre or f"Anexo de {unidad_display}",
        "esGenerico": es_generico or (not nombre and not es_vacante),
        "esVacante": es_vacante,
        "correos": correos,
        "anexo": str(anexo).strip() if anexo is not None else None,
        "cumpleanos": str(cumple).strip() if cumple else None,
        "seccion": seccion,
    }
    anexos_rows.append(row)
    for e in correos:
        anexos_por_correo.setdefault(e, row)

# ---------------------------------------------------------------------------
# 2) Pestaña "Dotación": planta oficial (RUT, grado, calidad jurídica, cargo)
# ---------------------------------------------------------------------------
ws = wb["Dotación"]
SECCION_DOTACION = {
    "ICA": "corte",
    "INSOLVENCIA": "insolvencia",
    "CINJ": "cinj",
    "CSMP": "csmp",
    "ULE": "ule",
}
UNIDAD_LABEL = {
    "ule": "Unidad de Liquidaciones Especializadas",
    "insolvencia": "Unidad de Insolvencia y Reemprendimiento",
    "cinj": "Centro Integrado de Notificaciones Judiciales",
    "csmp": "Centro de Seguimiento de Medidas de Protección",
}
dotacion_rows = []
for r in range(6, 99):
    rut = ws.cell(row=r, column=1).value
    nombre_raw = ws.cell(row=r, column=2).value
    grado = ws.cell(row=r, column=3).value
    calidad = ws.cell(row=r, column=4).value
    cargo = ws.cell(row=r, column=5).value
    correo_raw = ws.cell(row=r, column=6).value
    unidad_code = ws.cell(row=r, column=7).value
    if not nombre_raw:
        continue
    nombre, _ = clean_name(nombre_raw)
    correos = split_emails(correo_raw)
    seccion = SECCION_DOTACION.get(norm_key(unidad_code) if unidad_code else "", "corte")
    dotacion_rows.append(
        {
            "rut": str(rut).strip() if rut else None,
            "nombre": nombre,
            "grado": str(grado).strip() if grado else None,
            "calidadJuridica": title_case(str(calidad).strip()) if calidad else None,
            "cargo": title_case(str(cargo).strip()) if cargo else None,
            "correos": correos,
            "seccion": seccion,
        }
    )

# ---------------------------------------------------------------------------
# Merge Anexos + Dotación -> personas de Corte / Insolvencia / CINJ / CSMP / ULE
# ---------------------------------------------------------------------------
people: list[dict] = []
used_anexos_emails: set[str] = set()

for d in dotacion_rows:
    email = d["correos"][0] if d["correos"] else None
    a = anexos_por_correo.get(email) if email else None
    if a:
        used_anexos_emails.add(email)
    seccion = d["seccion"]
    if seccion in UNIDAD_LABEL:
        # Insolvencia / CINJ / CSMP / ULE son equipos pequeños y cohesionados:
        # se agrupan bajo una única etiqueta en vez de subdividir por el
        # rótulo (a veces inconsistente) de la hoja de Anexos.
        unidad = UNIDAD_LABEL[seccion]
    else:
        unidad = a["unidad"] if a else "Otros Cargos de Corte"
    # Dotación es la planta oficial (nombre legal completo, verificado con RUT);
    # se usa siempre que exista, para evitar erratas de tipeo del directorio de anexos.
    nombre = d["nombre"]
    correos = sort_emails(a["correos"]) if a and a["correos"] else d["correos"]
    people.append(
        {
            "id": None,
            "nombre": nombre,
            "cargo": d["cargo"],
            "unidad": unidad,
            "seccion": seccion,
            "tribunal": "Corte de Apelaciones de La Serena" if seccion == "corte" else UNIDAD_LABEL.get(seccion),
            "correos": correos,
            "anexo": a["anexo"] if a else None,
            "cumpleanos": a["cumpleanos"] if a else None,
            "rut": d["rut"],
            "grado": d["grado"],
            "calidadJuridica": d["calidadJuridica"],
            "esGenerico": False,
            "fuente": "dotacion+anexos" if a else "dotacion",
        }
    )

# Filas de Anexos sin contraparte en Dotación (correos generales, gendarmería,
# portería, periodistas suplentes, etc.)
for a in anexos_rows:
    email = a["correos"][0] if a["correos"] else None
    if email and email in used_anexos_emails:
        continue
    if not a["nombre"] and not a["correos"] and not a["anexo"]:
        continue
    seccion = a["seccion"]
    unidad = UNIDAD_LABEL[seccion] if seccion in UNIDAD_LABEL else a["unidad"]
    people.append(
        {
            "id": None,
            "nombre": a["nombre"] if not a["esVacante"] else "(Cargo vacante)",
            "cargo": None,
            "unidad": unidad,
            "seccion": seccion,
            "tribunal": "Corte de Apelaciones de La Serena" if seccion == "corte" else UNIDAD_LABEL.get(seccion),
            "correos": sort_emails(a["correos"]),
            "anexo": a["anexo"],
            "cumpleanos": a["cumpleanos"],
            "rut": None,
            "grado": None,
            "calidadJuridica": None,
            "esGenerico": a["esGenerico"],
            "vacante": a["esVacante"],
            "fuente": "anexos",
        }
    )

# ---------------------------------------------------------------------------
# 3) Pestaña "correos tribunales": ficha rápida de cada tribunal (switchboard)
# ---------------------------------------------------------------------------
ws = wb["correos tribunales"]
COMPETENCIA_COLS = list(range(7, 13))  # Civil..Reformado
COMPETENCIA_HEADERS = [str(ws.cell(row=2, column=c).value).strip() for c in COMPETENCIA_COLS]

tribunales_ficha: dict[str, dict] = {}
special_correo_general: dict[str, str] = {}
for r in range(3, 31):
    tribunal = ws.cell(row=r, column=4).value
    if not tribunal:
        continue
    tribunal = str(tribunal).strip()
    correo = ws.cell(row=r, column=6).value
    ministro = ws.cell(row=r, column=5).value
    telefono = ws.cell(row=r, column=13).value
    competencias = []
    for c, label in zip(COMPETENCIA_COLS, COMPETENCIA_HEADERS):
        v = ws.cell(row=r, column=c).value
        if v:
            competencias.append(label.strip())
    key = norm_key(tribunal)
    if "LIQUIDACIONES" in key or "NOTIFICACIONES JUDICIALES" in key or "MEDIDAS DE PROTECCION" in key:
        special_correo_general[key] = str(correo).strip() if correo else None
        continue
    tribunales_ficha[key] = {
        "nombre": tribunal,
        "correo": str(correo).strip() if correo else None,
        "telefono": str(telefono).strip() if telefono else None,
        "ministroVisitador": title_case(str(ministro).strip()) if ministro else None,
        "competencias": competencias,
    }

# ---------------------------------------------------------------------------
# 4) Pestaña "Jurisdicción": personal de los 26 tribunales + CAPJ Zonal +
#    Unidad de Apoyo a Tribunales Civiles (excluye lo ya cubierto por Corte,
#    CINJ, ULE y CSMP para no duplicar).
# ---------------------------------------------------------------------------
ws = wb["Jurisdicción"]
EXCLUIR_TRIBUNAL_KEYS = {
    norm_key("CORTE DE APELACIONES DE LA SERENA"),
}
EXCLUIR_CONTAINS = [
    "NOTIFICACIONES",
    "LIQUIDACIONES",
    "OBSERVACION Y CONTROL DE FAMILIA",
    "APOYO A TRIBUNALES CIVILES",  # mismo equipo que Dotación/INSOLVENCIA
]

MAPA_CORREOS_TRIBUNALES = {
    "1 JUZGADO DE LETRAS DE LA SERENA": "PRIMER JUZGADO DE LETRAS DE LA SERENA",
    "2 JUZGADO DE LETRAS DE LA SERENA": "SEGUNDO JUZGADO DE LETRAS DE LA SERENA",
    "3 JUZGADO DE LETRAS DE LA SERENA": "TERCER JUZGADO DE LETRAS DE LA SERENA",
    "1 JUZGADO DE LETRAS DE COQUIMBO": "PRIMER JUZGADO DE LETRAS DE COQUIMBO",
    "2 JUZGADO DE LETRAS DE COQUIMBO": "SEGUNDO JUZGADO DE LETRAS DE COQUIMBO",
    "3 JUZGADO LETRAS DE COQUIMBO": "TERCER JUZGADO DE LETRAS DE COQUIMBO",
    "1 JUZGADO DE LETRAS DE OVALLE": "PRIMER JUZGADO DE LETRAS DE OVALLE",
    "2 JUZGADO DE LETRAS DE OVALLE": "SEGUNDO JUZGADO DE LETRAS DE OVALLE",
    "3 JUZGADO LETRAS DE OVALLE": "TERCER JUZGADO DE LETRAS DE OVALLE",
    "JUZGADO DE LETRAS DE ILLAPEL": "JUZGADO DE LETRAS DE ILLAPEL",
    "JUZGADO DE LETRAS VICUNA": "JUZGADO DE LETRAS DE VICUÑA",
    "JUZGADO DE LETRAS DE ANDACOLLO": "JUZGADO DE LETRAS Y GARANTIA DE ANDACOLLO",
    "JUZGADO DE LETRAS Y GARANTIA DE COMBARBALA": "JUZGADO DE LETRAS Y GARANTIA DE COMBARBALA",
    "JUZGADO DE LETRAS Y GARANTIA DE LOS VILOS": "JUZGADO DE LETRAS Y GARANTIA DE LOS VILOS",
    "JUZGADO GARANTIA DE LA SERENA": "JUZGADO DE GARANTIA DE LA SERENA",
    "JUZGADO GARANTIA DE COQUIMBO": "JUZGADO DE GARANTIA DE COQUIMBO",
    "JUZGADO GARANTIA DE OVALLE": "JUZGADO DE GARANTIA DE OVALLE",
    "JUZGADO GARANTIA DE ILLAPEL": "JUZGADO DE GARANTIA DE ILLAPEL",
    "JUZGADO GARANTIA VICUNA": "JUZGADO DE GARANTIA DE VICUÑA",
    "JUZGADO DE LETRAS DEL TRABAJO DE LA SERENA": "JUZGADO DE LETRAS DEL TRABAJO DE LA SERENA",
    "JUZGADO FAMILIA LA SERENA": "JUZGADO DE FAMILIA DE LA SERENA",
    "JUZGADO FAMILIA COQUIMBO": "JUZGADO DE FAMILIA DE COQUIMBO",
    "JUZGADO FAMILIA DE OVALLE": "JUZGADO DE FAMILIA DE OVALLE",
    "TRIBUNAL DE JUICIO ORAL EN LO PENAL DE LA SERENA": "TRIBUNAL DE JUICIO ORAL EN LO PENAL DE LA SERENA",
    "TRIBUNAL DE JUICIO ORAL EN LO PENAL DE OVALLE": "TRIBUNAL DE JUICIO ORAL EN LO PENAL DE OVALLE",
}


def display_tribunal_name(raw: str) -> str:
    target = MAPA_CORREOS_TRIBUNALES.get(norm_key(raw))
    if target:
        ficha = tribunales_ficha.get(norm_key(target))
        if ficha:
            return ficha["nombre"]
    raw = str(raw).strip()
    # Los nombres ya bien capitalizados en la fuente (p.ej. "CAPJ Zonal La Serena")
    # se respetan tal cual; solo se re-capitalizan los que vienen en MAYÚSCULAS.
    if raw == raw.upper():
        return title_case(raw)
    return raw


jurisdiccion_rows = []
vacantes = 0
for r in range(2, 675):
    unidad_col = ws.cell(row=r, column=2).value
    tribunal_raw = ws.cell(row=r, column=10).value
    if not tribunal_raw:
        continue
    tkey = norm_key(tribunal_raw)
    if tkey in EXCLUIR_TRIBUNAL_KEYS:
        continue
    if any(x in tkey for x in EXCLUIR_CONTAINS):
        continue
    nombre_raw = ws.cell(row=r, column=7).value
    correo_raw = ws.cell(row=r, column=8).value
    cargo = ws.cell(row=r, column=9).value
    estado = ws.cell(row=r, column=11).value
    calidad = ws.cell(row=r, column=12).value
    suplente = ws.cell(row=r, column=14).value
    estado_norm = norm_key(estado) if estado else ""
    if estado_norm == "VACANTE":
        vacantes += 1
    if not nombre_raw and estado_norm != "VACANTE":
        continue
    nombre, _ = clean_name(nombre_raw) if nombre_raw else ("(Cargo vacante)", False)
    correos = split_emails(correo_raw)
    tribunal_display = display_tribunal_name(tribunal_raw)
    seccion = "capj" if "CAPJ" in tkey else "tribunal"
    jurisdiccion_rows.append(
        {
            "id": None,
            "nombre": nombre,
            "cargo": title_case(str(cargo).strip()) if cargo else None,
            "unidad": tribunal_display,
            "seccion": seccion,
            "tribunal": tribunal_display,
            "correos": correos,
            "anexo": None,
            "cumpleanos": None,
            "rut": None,
            "grado": None,
            "calidadJuridica": title_case(str(calidad).strip()) if calidad else None,
            "esGenerico": False,
            "vacante": estado_norm == "VACANTE",
            "suplente": title_case(str(suplente).strip()) if suplente else None,
            "fuente": "jurisdiccion",
        }
    )

people.extend(jurisdiccion_rows)

# Adjuntar correo general / teléfono a las fichas ule/cinj/csmp usando la
# hoja "correos tribunales" (filas de unidades especiales)
UNIDAD_KEY_ESPECIAL = {
    "ule": "LIQUIDACIONES",
    "cinj": "NOTIFICACIONES JUDICIALES",
    "csmp": "MEDIDAS DE PROTECCION",
}
correo_general_seccion = {}
for seccion, needle in UNIDAD_KEY_ESPECIAL.items():
    for k, correo in special_correo_general.items():
        if needle in k:
            correo_general_seccion[seccion] = correo

# ---------------------------------------------------------------------------
# Adjuntar id único + ficha de tribunal a cada persona
# ---------------------------------------------------------------------------
slug_count: dict[str, int] = {}
for p in people:
    base = slugify(f"{p['nombre']}-{p.get('unidad') or ''}")
    n = slug_count.get(base, 0)
    slug_count[base] = n + 1
    p["id"] = base if n == 0 else f"{base}-{n}"
    ficha = tribunales_ficha.get(norm_key(p["tribunal"])) if p.get("tribunal") else None
    p["fichaTribunal"] = ficha

tribunales_list = sorted(tribunales_ficha.values(), key=lambda t: t["nombre"])
for t in tribunales_list:
    t["id"] = slugify(t["nombre"])
    t["comuna"] = extraer_comuna(t["nombre"])

for p in people:
    if p["seccion"] in {"corte", "insolvencia", "cinj", "csmp", "ule", "capj"}:
        p["comuna"] = "La Serena"
    else:
        p["comuna"] = extraer_comuna(p["unidad"]) if p.get("unidad") else None

# ---------------------------------------------------------------------------
# Protección de datos personales: la plataforma es de uso interno, por lo que
# no se publica el RUT (identificador legal sensible, sin uso en la UI) ni
# correos personales (Gmail/Hotmail/etc.) — solo direcciones institucionales
# @pjud.cl.
# ---------------------------------------------------------------------------
DOMINIOS_INSTITUCIONALES = ("@pjud.cl", "@poderjudicial.cl")
personas_sin_correo = 0
for p in people:
    p.pop("rut", None)
    institucionales = [e for e in p["correos"] if e.endswith(DOMINIOS_INSTITUCIONALES)]
    p["correos"] = institucionales
    if not institucionales:
        personas_sin_correo += 1

data = {
    "generatedAt": "2026-07-10",
    "totalPersonas": len(people),
    "people": people,
    "tribunales": tribunales_list,
    "correoGeneralSeccion": correo_general_seccion,
}

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

print("Personas totales:", len(people))
print("Vacantes detectadas (excluidas del listado si sin nombre):", vacantes)
from collections import Counter
print(Counter(p["seccion"] for p in people))
print("Tribunales con ficha:", len(tribunales_list))
print("Correo general por sección especial:", correo_general_seccion)
print("Personas sin correo institucional tras el filtro de privacidad:", personas_sin_correo)
